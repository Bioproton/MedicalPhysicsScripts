import numpy as np
import os
from openpyxl import load_workbook
import random
from NOVO_to_FLUKA_coordinates import NOVO_to_FLUKA_coordinate_transform

# Read Excel-file, gather nescessary information

# First initialize script, point to where the information is located

# Structure of the detector along with corresponding names
layers = {
    "Layer1": ["10", "11"],
    "Layer2": ["20", "21"],
    "Layer3": ["30", "31", "32", "33"],
    "Layer4": ["40", "41", "42", "43"]
}

# Conversion between naming schemes "bar1 - bar14" and "10 - 43"
scint_convert = [
    "10", "11", 
    "20", "21", "50", "51", 
    "30", "31", "32", "33", 
    "40", "41", "42", "43"
]

def calc_scintillator_coordinates(Worksheet, Layer, Scintillator):

        # Template for coordinates that need to be calculated for a single bar [need 14 of these]
        bar = { 
            "bar" : [0, 0, 0, 0, 0, 0],
            "HOut": [0, 0, 0, 0, 0, 0],
            "HIn" : [0, 0, 0, 0, 0, 0],
            "Shm" : [0, 0, 0],
            "SHolp": [0, 0, 0, 0, 0, 0],
            "SHoln": [0, 0, 0, 0, 0, 0],
            "CHol": [0, 0, 0],
            "Clpos": [0, 0, 0, 0, 0, 0, 0],
            "Clneg": [0, 0, 0, 0, 0, 0, 0],
            "Glpos": [0, 0, 0, 0, 0, 0, 0],
            "Glneg": [0, 0, 0, 0, 0, 0, 0],
            "Void": [0, 0, 0, 0, 0, 0]
        }   

        # Template for coordinates that need to be calculated for a single bar [only need 4 of these]
        plane = {
            "Hpos": 0,
            "Hneg": 0,
            "S1pos": 0,
            "S1neg": 0,
            "S2pos": 0,
            "S2neg": 0,
            "ELpos": 0,
            "ELneg": 0,
            "SbO_p": [0, 0, 0, 0, 0, 0],    # Outer support bar, positive side
            "SbI_p": [0, 0, 0, 0, 0, 0],    # Inner support bar, positive side
            "SbO_n": [0, 0, 0, 0, 0, 0],    # Outer support bar, negative side
            "SbI_n": [0, 0, 0, 0, 0, 0]     # Inner support bar, negative side
        }

        # Conversion between naming schemes "bar1 - bar14" and "10 - 43"
        scint_convert = [
            "10", "11", 
            "20", "21", "50", "51", 
            "30", "31", "32", "33", 
            "40", "41", "42", "43"
        ]

        # Figuring out which row to get the scintillator bar coordinates:
        rowN = 1 + 9 + scint_convert.index(Scintillator)

        # Inserting bar coordinates
        bar["bar"] = [round(cell.value, 5) for row in Worksheet.iter_rows(min_row = rowN, max_row=rowN, min_col=21, max_col=26) for cell in row]


        # Splitting up the calculations based on bar orientation
        # -------------------------ODD LAYERS---------------------------------
        if Layer in ["Layer1", "Layer3"]:

            # Finding layer column to use for casing information
            if Layer == "Layer1":
                columnN = "K"

                # Support columns for layer 1 bar

                # Outer support bar (right/positive)
                plane["SbO_p"] = [
                    round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value, 2), 
                    round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(29)].value, 2),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(29)].value, 2)), 
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2))
                ]

                # Inner support bar (right/positive)
                plane["SbI_p"] = [
                    round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(35)].value, 2), 
                    round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(29)].value - 0.1 * Worksheet[columnN + str(35)].value, 2),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)),
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)), 
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(35)].value, 2))
                ]

                # Outer support bar (left/negative)
                plane["SbO_n"] = [
                    round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(29)].value, 2),
                    round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value, 2), 
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(29)].value, 2)), 
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2)) 
                ]

                # Inner support bar (left/negative)
                plane["SbI_n"] = [
                    round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2),
                    round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(35)].value, 2), 
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)),
                    float(round(np.mean([Worksheet["W" + str(10)].value, Worksheet["X" + str(11)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)),
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)), 
                    float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value - 0.1 * Worksheet[columnN + str(35)].value, 2))
                ]

            elif Layer == "Layer3": 
                columnN = "Q"
            
            # Calcuating outer casing box
            bar["HOut"] = [
                round(Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(30)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(30)].value, 2)
            ]

            # Calcuating inner casing box
            bar["HIn"] = [
                round(Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)             
            ]

            # Calculating shim cylinder
            bar["Shm"] = [
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(38)].value, 2)
            ]

            # Calculating square hole for shim 1 positive X
            bar["SHolp"] = [
                round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating square hole for shim 1 negative X
            bar["SHoln"] = [
                round(Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value, 2),
                round(Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating round hole for shim 2 positive and negative X
            bar["CHol"] = [
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(44)].value, 2)
            ]
            
            # Calculating cylinder (readout thing) for positive X
            bar["Clpos"] = [
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value, 2),
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                0,
                round(0.1 * Worksheet[columnN + str(47)].value, 2)
            ]

            # Calculating cylinder (readout thing) for negative X
            bar["Clneg"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value, 2),
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(-0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                0,
                round(0.1 * Worksheet[columnN + str(47)].value, 2)
            ]
        
            # Calculating void box box
            bar["Void"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * Worksheet[columnN + str(46)].value, 2), 
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * Worksheet[columnN + str(46)].value, 2), 
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2)),
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]) + 0.1 * Worksheet[columnN + str(38)].value, 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2)), 
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) + 0.1 * Worksheet[columnN + str(38)].value, 2))
            ]

            # Calculating cylinder (readout thing) for positive X
            bar["Glpos"] = [
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * Worksheet[columnN + str(49)].value, 2),
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                0,
                round(0.1 * Worksheet[columnN + str(47)].value - 0.1 * Worksheet[columnN + str(49)].value, 2)
            ]

            # Calculating glass portion of PMT for negative X
            bar["Glneg"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * Worksheet[columnN + str(49)].value, 2),
                float(round(np.mean([Worksheet["W" + str(rowN)].value, Worksheet["X" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                round(-0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                0,
                round(0.1 * Worksheet[columnN + str(47)].value - 0.1 * Worksheet[columnN + str(49)].value, 2)
            ]

        
            # Calculating plane coordinates
            
            # House edges
            plane["Hpos"] = round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2)
            plane["Hneg"] = round((Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value), 2)

            # Shim 1 edges
            plane["S1pos"] = round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value, 2)
            plane["S1neg"] = round((Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value), 2)

            # Shim 2 edges
            plane["S2pos"] = round(Worksheet["V" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value + 0.1 * Worksheet[columnN + str(42)].value, 2)
            plane["S2neg"] = round((Worksheet["U" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value - 0.1 * Worksheet[columnN + str(42)].value), 2)

            # PMT glass confinements
            plane["ELpos"] = round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * 0.5 * Worksheet[columnN + str(46)].value, 2)
            plane["ELneg"] = round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * 0.5 * Worksheet[columnN + str(46)].value, 2)

        # -------------------------EVEN LAYERS---------------------------------
        elif Layer == "Layer4":
            
            # Saving which column the casing information is stored in
            columnN = "Q"

            # Calcuating outer casing box
            bar["HOut"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value- 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value+ 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(30)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(30)].value, 2)
            ]

            # Calcuating inner casing box
            bar["HIn"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)              
            ]

            # Calculating shim cylinder
            bar["Shm"] = [
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(38)].value, 2)
            ]

            # Calculating square hole for shim 1 positive Y
            bar["SHolp"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating square hole for shim 1 negative Y
            bar["SHoln"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating round hole for shim 2 positive and negative Y
            bar["CHol"] = [
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(0.1 * Worksheet[columnN + str(44)].value, 2)
            ]
            
            # Calculating cylinder (readout thing) for positive Y
            bar["Clpos"] = [
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value, 2),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                0,
                round(0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                round(0.1 * Worksheet[columnN + str(47)].value, 2)
            ]

            # Calculating cylinder (readout thing) for negative Y
            bar["Clneg"] = [
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value, 2),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                0,
                round(-0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                round(0.1 * Worksheet[columnN + str(47)].value, 2)
            ]
        
            # Calculating void box box
            bar["Void"] = [
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2)),
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]) + 0.1 * Worksheet[columnN + str(38)].value, 2)),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * Worksheet[columnN + str(46)].value, 2), 
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * Worksheet[columnN + str(46)].value, 2),  
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) - 0.1 * Worksheet[columnN + str(38)].value, 2)), 
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]) + 0.1 * Worksheet[columnN + str(38)].value, 2))
            ]

            # Calculating cylinder (readout thing) for positive Y
            bar["Glpos"] = [
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * Worksheet[columnN + str(49)].value, 2),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                0,
                round(0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                round(0.1 * Worksheet[columnN + str(47)].value - 0.1 * Worksheet[columnN + str(49)].value, 2)
            ]

            # Calculating cylinder (readout thing) for negative Y
            bar["Glneg"] = [
                float(round(np.mean([Worksheet["U" + str(rowN)].value, Worksheet["V" + str(rowN)].value]), 2)),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * Worksheet[columnN + str(49)].value, 2),
                float(round(np.mean([Worksheet["Y" + str(rowN)].value, Worksheet["Z" + str(rowN)].value]), 2)),
                0,
                round(-0.1 * Worksheet[columnN + str(46)].value, 2),
                0,
                round(0.1 * Worksheet[columnN + str(47)].value - 0.1 * Worksheet[columnN + str(49)].value, 2)
            ]

            # Calculating plane coordinates
            
            # House edges
            plane["Hpos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2)
            plane["Hneg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2)

            # Shim 1 edges
            plane["S1pos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value, 2)
            plane["S1neg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value, 2)

            # Shim 2 edges
            plane["S2pos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(37)].value + 0.1 * Worksheet[columnN + str(42)].value, 2)
            plane["S2neg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(37)].value - 0.1 * Worksheet[columnN + str(42)].value, 2)

            # PMT glass confinements
            plane["ELpos"] = round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(48)].value + 0.1 * 0.5 * Worksheet[columnN + str(46)].value, 2)
            plane["ELneg"] = round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(48)].value - 0.1 * 0.5 * Worksheet[columnN + str(46)].value, 2)
        
        # Layer 2 (U301) needs special treatment
        elif Layer == "Layer2":

            # Saving which column the casing information is stored in
            columnN = "W"

            # Support columns for layer 2 bar

            # Outer support bar (right/positive)
            plane["SbO_p"] = [
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(29)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(29)].value, 2), 
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2)
            ]

            # Inner support bar (right/positive)
            plane["SbI_p"] = [
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)),
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(35)].value, 2),
                round(Worksheet["X" + str(rowN)].value + 0.1 * Worksheet[columnN + str(45)].value + 0.1 * Worksheet[columnN + str(29)].value - 0.1 * Worksheet[columnN + str(35)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2), 
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)
            ]

            # Outer support bar (left/negative)
            plane["SbO_n"] = [
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value, 2)),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(29)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(29)].value, 2), 
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2)
            ]

            # Inner support bar (left/negative)
            plane["SbI_n"] = [
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) - 0.1 * 0.5 * Worksheet[columnN + str(28)].value + 0.1 * Worksheet[columnN + str(35)].value, 2)),
                float(round(np.mean([Worksheet["U" + str(12)].value, Worksheet["V" + str(13)].value]) + 0.1 * 0.5 * Worksheet[columnN + str(28)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2),
                round(Worksheet["W" + str(rowN)].value - 0.1 * Worksheet[columnN + str(45)].value - 0.1 * Worksheet[columnN + str(35)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(29)].value + 0.1 * Worksheet[columnN + str(35)].value, 2), 
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value - 0.1 * Worksheet[columnN + str(35)].value, 2)
            ]

            # Calcuating outer casing box
            bar["HOut"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value- 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value+ 0.1 * Worksheet[columnN + str(30)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(30)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(30)].value, 2)
            ]

            # Calcuating inner casing box
            bar["HIn"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2), 
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)             
            ]

            # Calculating shim BOX (not cylinder)
            bar["Shm"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(44)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(44)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value, 2)
            ]
    
            # Calculating square hole for shim 1 positive Y
            bar["SHolp"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(36)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating square hole for shim 1 negative Y
            bar["SHoln"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(36)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value, 2)
            ]

            # Calculating square hole for shim 2 positive and negative Y
            bar["CHol"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value + 0.1 * Worksheet[columnN + str(36)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value - 0.1 * Worksheet[columnN + str(36)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(44)].value, 2),
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(44)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value + 0.1 * Worksheet[columnN + str(36)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value - 0.1 * Worksheet[columnN + str(36)].value, 2)
            ]

            # Calculating void box box
            bar["Void"] = [
                round(Worksheet["U" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2),
                round(Worksheet["V" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value, 2),
                round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(44)].value, 2), 
                round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(44)].value, 2),
                round(Worksheet["Y" + str(rowN)].value - 0.1 * Worksheet[columnN + str(33)].value - 0.1 * Worksheet[columnN + str(39)].value, 2),
                round(Worksheet["Z" + str(rowN)].value + 0.1 * Worksheet[columnN + str(33)].value + 0.1 * Worksheet[columnN + str(40)].value, 2)
            ]
            
            # Calculating plane coordinates
            
            # House edges
            plane["Hpos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value, 2)
            plane["Hneg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value, 2)

            # Shim 1 edges
            plane["S1pos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(36)].value, 2)
            plane["S1neg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(36)].value, 2)

            # Shim 2 edges
            plane["S2pos"] = round(Worksheet["X" + str(rowN)].value - 0.1 * Worksheet[columnN + str(34)].value + 0.1 * Worksheet[columnN + str(36)].value + 0.1 * Worksheet[columnN + str(44)].value, 2)
            plane["S2neg"] = round(Worksheet["W" + str(rowN)].value + 0.1 * Worksheet[columnN + str(34)].value - 0.1 * Worksheet[columnN + str(36)].value - 0.1 * Worksheet[columnN + str(44)].value, 2)

        return bar, plane


# Function for creating/calculating coordinates:
# For each layer, calc all the plane coordinates
# For each scintillator in the layer, calc the rest of geometry coordinates
def load_all_coordinates(excel_file):

    scintillator_coordinate_dict = {
        "10" : None, "11" : None, 
        "20" : None, "21" : None,
        "30" : None, "31" : None, "32" : None, "33" : None, 
        "40" : None, "41" : None, "42" : None, "43" : None
    }

    plane_coordinate_dict = {
        "Layer1": None, "Layer2" : None, "Layer3" : None, "Layer4" : None
    }

    # Structure of the detector along with corresponding names
    layers = {
        "Layer1": ["10", "11"],
        "Layer2": ["20", "21"],
        "Layer3": ["30", "31", "32", "33"],
        "Layer4": ["40", "41", "42", "43"]
    }

    # Load the workbook
    workbook = load_workbook(excel_file, data_only=True)
    worksheet = workbook["Main"]

    for layer, scintillators in layers.items():
        for scintillator in scintillators:
            #print(f"{layer}, scintillator {scintillator}")

            # Calculate scintillator and plane coordinates
            scintillator_coordinates, plane_coordinates = calc_scintillator_coordinates(
                Worksheet=worksheet, Layer=layer, Scintillator=scintillator
            )

            # Save scintillator/casing coordinates for current scintillator
            scintillator_coordinate_dict[scintillator] = scintillator_coordinates
            
        # Save plane coordinates for current layer
        plane_coordinate_dict[layer] = plane_coordinates

    return scintillator_coordinate_dict, plane_coordinate_dict


def create_fluka_input(inp_file_path, excel_file, distance_PMMA_detec = 54.2, distance_beam_PMMA = 46.6, inp_file_name = None, pmma_phantom=True):
    # Creates a full .inp file with the PTB model in FLUKA that can be used with FLUKA simulation
    # Credits to Lars Fredrik Fjæra/Helge Henjum for the .inp format from sort_dicoms_impt.py

    # Inputs:
    #   - inp_file_path: Path where one wishes to save the .inp file
    #   - excel_file: File path to Excel-sheet with detector position informatin
    #   - distance_PMMA_detec: Distance from PMMA phantom center to first layer scintillator center [z in detector frame, x in FLUKA viewport]
    #   - distance_beam_PMMA: Distance from beam exit to PMMA phantom surface [x in FLUKA start_translat, z in FLUKA viewport]
    #   - inp_file_name: Default: "OncoRay_model.inp"
    #   - pmma_phantom: Flag to check whether to include the cylindrical PMMA-phantom or not: Default: True


    # Structure of the detector along with corresponding names
    layers = {
        "Layer1": ["10", "11"],
        "Layer2": ["20", "21"],
        "Layer3": ["30", "31", "32", "33"],
        "Layer4": ["40", "41", "42", "43"]
    }

    # Conversion between naming schemes "bar1 - bar14" and "10 - 43"
    scint_convert = [
        "10", "11", 
        "20", "21", 
        "30", "31", "32", "33", 
        "40", "41", "42", "43"
    ]

    nominal_beam_exit_PMMA_distance = 46.6

    def list_to_str(a_list):
        # Simple function to convert list of floats to string

        # If there is only one element in the list, return the string
        if len(a_list) == 1:
            return str(a_list[0])
        
        output = ""
        for value in a_list:
            output += " " + str(value)

        return output[1:]

    # Check if all coordinates have been calculated. If not, calculate them again
    scintillator_coordinate_dict, plane_coordinate_dict = load_all_coordinates(excel_file=excel_file)

    # Assign default .inp file name if no name is given
    if inp_file_name is None:
        inp_file_name = "OncoRay_model.inp"
    else:
        if inp_file_name[-4:] != ".inp":
            inp_file_name += ".inp"
    
    # If .inp file already exists, create a copy of it to avoid overriding a previous version
    n_copy = 0
    while os.path.exists(inp_file_path + r"\\" + inp_file_name):
        n_copy += 1
        inp_file_name = "OncoRay_model(" + str(n_copy) + ").inp"

    f = open(inp_file_path + r"\\" + inp_file_name, "w")

    # -----------------------START OF FLUKA INPUT-----------------------

    # Correct format template for FLUKA/FLAIR card. All inputs must be strings!
    # f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("","","","","","","",""))
    f.write("* OncoRay model with scintillator casing details")
    

    # --------------------DEFAULTS, BEAM AND ROTATION-------------------
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("TITLE","","","","","","","OLDFLAIR"))

    f.write("\n* Simulation defaults plus physics cards that allows Coalescense and Evaporation (for PGs)")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("DEFAULTS","","","","","","","PRECISIO"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("PHYSICS","3.0","","","","","","EVAPORAT"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("PHYSICS","1.0","","","","","","COALESCE"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("PHYSICS","1.0","0.005","0.15","2.0","2.0","2.","IONSPLIT"))

    f.write("\n* Beam characteristics and beam position")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("BEAM","-0.190","0.0","0.0","-1.24","-0.85","0.0","PROTON"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("BEAMPOS","0.0","0.0","-47.2","0.0","0.0","",""))

    f.write("\n* NOVCoDA rotation cards. Adjust azimuthal angles for rotations")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ROT-DEFI","300.","0.0","90.0","0.0","0.0","-10.6","NOVO_rot"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ROT-DEFI","100.","0.0","90.0","0.0","0.0","0.0","NOVO_rot"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ROT-DEFI","300.","0.0","90.0","0.0","0.0","0.0","NOVO_rot"))

    # ------------------------------GEOMETRY----------------------------
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("GEOBEGIN","","","","","","","COMBNAME"))
    f.write("\n    0    0")

    # Target, void and blackbody
    f.write("\n* Black body")
    f.write("\nSPH blkbody    0.0 0.0 0.0 100000.0") #Free format card
    f.write("\n* Void sphere")
    f.write("\nSPH Void       0.0 0.0 0.0 10000.0") #Free format card
    f.write("\n* Air subsphere")
    f.write("\nSPH Air       0.0 0.0 0.0 1000.0") #Free format card

    # -------PMMA target-------
    f.write(f"\n$start_translat  0.0 0.0 {str(round(distance_beam_PMMA - nominal_beam_exit_PMMA_distance, 2))}")
    f.write(f"\nRCC casing    0.0 0.0 0.0 0.0 0.0 40.0 7.5")

    # First half of inserts (1.1 - 1.10)
    for i in range(10):
        f.write(f"\nRCC ins1.{i + 1}    0.0 0.0 {str(i * 2.0)} 0.0 0.0 2.0 2.45")
    # Second half of inserts (2.1 - 2.10)
    for i in range(10):
        f.write(f"\nRCC ins2.{i + 1}    0.0 0.0 {str(i * 2.0 + 20.0)} 0.0 0.0 2.0 2.45")
    f.write(f"\n$end_translat")


    # -------PTB model-------
    f.write("\n$start_transform NOVO_rot") #Free format card
    f.write("\n*Detector positioning: [Horizontal, Vertical, Radial]")
    f.write(f"\n$start_translat 20.0 0.0 {str(round(distance_PMMA_detec - 0.6 + 10.6, 2))}") #Free format card 
    # Default distance radial: 54.2 cm - 0.6 cm + 10.6 cm = 64.2 cm

    # Format: f.write("\nGEO NAME   COORD")
    for layer, scintillators in layers.items():

        # For odd layers (PMT1 and PMT2)
        if layer == "Layer1" or layer == "Layer3":

            # Support bars for layer 1 only
            if layer == "Layer1":
                # Writing support bar information
                f.write(f"\n*Support bars for layer {layer[-1]}")
                f.write(f"\nRPP SbO_p{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbO_p"])}")
                f.write(f"\nRPP SbI_p{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbI_p"])}")
                f.write(f"\nRPP SbO_n{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbO_n"])}")
                f.write(f"\nRPP SbI_n{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbI_n"])}")
            
            f.write(f"\n*{layer}")
            # Writing plane information (once per layer)
            f.write(f"\nYZP Hpos{layer[-1]}   {str(plane_coordinate_dict[layer]["Hpos"])}")
            f.write(f"\nYZP Hneg{layer[-1]}   {str(plane_coordinate_dict[layer]["Hneg"])}")
            f.write(f"\nYZP S1pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S1pos"])}")
            f.write(f"\nYZP S1neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S1neg"])}")
            f.write(f"\nYZP S2pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S2pos"])}")
            f.write(f"\nYZP S2neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S2neg"])}")
            f.write(f"\nYZP ELpos{layer[-1]}   {str(plane_coordinate_dict[layer]["ELpos"])}")
            f.write(f"\nYZP ELneg{layer[-1]}   {str(plane_coordinate_dict[layer]["ELneg"])}")

            for scintillator in scintillators:
                f.write(f"\n*Scintillator {scintillator}")
                # Writing scintillator information
                f.write(f"\nRPP bar{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["bar"])}")
                f.write(f"\nRPP HOut{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HOut"])}")
                f.write(f"\nRPP HIn{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HIn"])}")
                f.write(f"\nXCC Shm{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Shm"])}")
                f.write(f"\nRPP SHolp{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHolp"])}")
                f.write(f"\nRPP SHoln{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHoln"])}")
                f.write(f"\nXCC CHol{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["CHol"])}")
                f.write(f"\nRCC Clpos{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Clpos"])}")
                f.write(f"\nRCC Clneg{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Clneg"])}")
                f.write(f"\nRPP Void{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Void"])}")
                f.write(f"\nRCC Glpos{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Glpos"])}")
                f.write(f"\nRCC Glneg{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Glneg"])}")
                
        # For even layer (U301)
        if layer == "Layer2":

            f.write(f"\n*{layer}")
            # Writing plane information (once per layer)
            f.write(f"\nXZP Hpos{layer[-1]}   {str(plane_coordinate_dict[layer]["Hpos"])}")
            f.write(f"\nXZP Hneg{layer[-1]}   {str(plane_coordinate_dict[layer]["Hneg"])}")
            f.write(f"\nXZP S1pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S1pos"])}")
            f.write(f"\nXZP S1neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S1neg"])}")
            f.write(f"\nXZP S2pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S2pos"])}")
            f.write(f"\nXZP S2neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S2neg"])}")

            # Writing support bar information
            f.write(f"\n*Support bars for layer {layer[-1]}")
            f.write(f"\nRPP SbO_p{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbO_p"])}")
            f.write(f"\nRPP SbI_p{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbI_p"])}")
            f.write(f"\nRPP SbO_n{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbO_n"])}")
            f.write(f"\nRPP SbI_n{layer[-1]}   {list_to_str(plane_coordinate_dict[layer]["SbI_n"])}")

            for scintillator in scintillators:
                f.write(f"\n*Scintillator {scintillator}")
                # Writing scintillator information
                f.write(f"\nRPP bar{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["bar"])}")
                f.write(f"\nRPP HOut{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HOut"])}")
                f.write(f"\nRPP HIn{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HIn"])}")
                f.write(f"\nRPP Shm{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Shm"])}")
                f.write(f"\nRPP SHolp{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHolp"])}")
                f.write(f"\nRPP SHoln{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHoln"])}")
                f.write(f"\nRPP CHol{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["CHol"])}")
                f.write(f"\nRPP Void{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Void"])}")
        
        # For even layer (PMT2)
        if layer == "Layer4":

            f.write(f"\n*{layer}")
            # Writing plane information (once per layer)
            f.write(f"\nXZP Hpos{layer[-1]}   {str(plane_coordinate_dict[layer]["Hpos"])}")
            f.write(f"\nXZP Hneg{layer[-1]}   {str(plane_coordinate_dict[layer]["Hneg"])}")
            f.write(f"\nXZP S1pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S1pos"])}")
            f.write(f"\nXZP S1neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S1neg"])}")
            f.write(f"\nXZP S2pos{layer[-1]}   {str(plane_coordinate_dict[layer]["S2pos"])}")
            f.write(f"\nXZP S2neg{layer[-1]}   {str(plane_coordinate_dict[layer]["S2neg"])}")
            f.write(f"\nXZP ELpos{layer[-1]}   {str(plane_coordinate_dict[layer]["ELpos"])}")
            f.write(f"\nXZP ELneg{layer[-1]}   {str(plane_coordinate_dict[layer]["ELneg"])}")

            for scintillator in scintillators:
                f.write(f"\n*Scintillator {scintillator}")
                # Writing scintillator information
                f.write(f"\nRPP bar{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["bar"])}")
                f.write(f"\nRPP HOut{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HOut"])}")
                f.write(f"\nRPP HIn{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["HIn"])}")
                f.write(f"\nYCC Shm{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Shm"])}")
                f.write(f"\nRPP SHolp{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHolp"])}")
                f.write(f"\nRPP SHoln{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["SHoln"])}")
                f.write(f"\nYCC CHol{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["CHol"])}")
                f.write(f"\nRCC Clpos{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Clpos"])}")
                f.write(f"\nRCC Clneg{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Clneg"])}")
                f.write(f"\nRPP Void{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Void"])}")
                f.write(f"\nRCC Glpos{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Glpos"])}")
                f.write(f"\nRCC Glneg{scintillator}   {list_to_str(scintillator_coordinate_dict[scintillator]["Glneg"])}")
   
    f.write("\n$end_translat") #Free format card
    f.write("\n$end_transform") #Free format card

    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("END","","","","","","",""))
    # -------------------------------REGION-----------------------------

    # Scintillator bars (sensitive to region definition order)
    f.write("\n* Scintillator regions: MREG = 1 - 14")  
    f.write("\nBAR10       5 +bar10")
    f.write("\nBAR11       5 +bar11")
    f.write("\nBAR20       5 +bar20")
    f.write("\nBAR21       5 +bar21")
    f.write("\nBAR30       5 +bar30")
    f.write("\nBAR31       5 +bar31")
    f.write("\nBAR32       5 +bar32")
    f.write("\nBAR33       5 +bar33")
    f.write("\nBAR40       5 +bar40")
    f.write("\nBAR41       5 +bar41")
    f.write("\nBAR42       5 +bar42")
    f.write("\nBAR43       5 +bar43")

    # -------PMMA phantom-------
    f.write("\n* PMMA phantom target")
    f.write("\nCASING       5 +casing " + 
            "-ins1.1 -ins1.2 -ins1.3 -ins1.4 -ins1.5 -ins1.6 -ins1.7 -ins1.8 -ins1.9 -ins1.10" + 
            "-ins2.1 -ins2.2 -ins2.3 -ins2.4 -ins2.5 -ins2.6 -ins2.7 -ins2.8 -ins2.9 -ins2.10"
        )
    
    for i in range(10):
        f.write(f"\nINS1.{i + 1}       5 +ins1.{i + 1}")
    for i in range(10):
        f.write(f"\nINS2.{i + 1}       5 +ins2.{i + 1}")

    f.write("\nAIR       5 +Air -Void10 -Void11 -Void20 -Void21 -Void30 -Void31 -Void32 -Void33 -Void40 -Void41 -Void42 -Void43 -casing -SbO_p1 -SbO_n1 - SbO_p2 - SbO_n2")
    f.write("\nVOID      5 +Void -Air")
    f.write("\nBLKBODY      5 +blkbody -Void")

    for layer, scintillators in layers.items():
        if layer == "Layer1" or layer == "Layer2":
            f.write(f"\n*Support bars for layer {layer[-1]}")
            f.write(f"\nSBO_P{layer[-1]}      5 +SbO_p{layer[-1]} -SbI_p{layer[-1]}")
            f.write(f"\nSBI_P{layer[-1]}      5 +SbI_p{layer[-1]}")
            f.write(f"\nSBO_N{layer[-1]}      5 +SbO_n{layer[-1]} -SbI_n{layer[-1]}")
            f.write(f"\nSBI_N{layer[-1]}      5 +SbI_n{layer[-1]}")

        for scintillator in scintillators:
            f.write(f"\n*Scintillator {scintillator}")
            # Scintillator definitions (including void...)
            f.write(f"\nHOUSE{scintillator}      5 +HOut{scintillator} -HIn{scintillator} -Hneg{layer[-1]} +Hpos{layer[-1]}")
            f.write(f"\nS1POS{scintillator}      5 +Shm{scintillator} -SHolp{scintillator} -Hpos{layer[-1]} +S1pos{layer[-1]}")
            f.write(f"\nS1NEG{scintillator}      5 +Shm{scintillator} -SHoln{scintillator} +Hneg{layer[-1]} -S1neg{layer[-1]}")
            f.write(f"\nS2POS{scintillator}      5 +Shm{scintillator} -CHol{scintillator} -S1pos{layer[-1]} +S2pos{layer[-1]}")
            f.write(f"\nS2NEG{scintillator}      5 +Shm{scintillator} -CHol{scintillator} +S1neg{layer[-1]} -S2neg{layer[-1]}")

            # Special case for Layer 2
            if layer == "Layer2":
                f.write(f"\nVOID{scintillator}      5" + 
                    f" +HIn{scintillator} -bar{scintillator} -Hneg{layer[-1]} +Hpos{layer[-1]}" + 
                    f"|+Void{scintillator} -HOut{scintillator} -Hneg{layer[-1]} +Hpos{layer[-1]}" +
                    f"|+SHolp{scintillator} -bar{scintillator} -Hpos{layer[-1]} +S1pos{layer[-1]}" +
                    f"|+CHol{scintillator} -S1pos{layer[-1]} +S2pos{layer[-1]}" +
                    f"|Void{scintillator} -Shm{scintillator} -Hpos{layer[-1]} + S2pos{layer[-1]}" +
                    f"|Void{scintillator} -S2pos{layer[-1]}" +
                    f"|SHoln{scintillator} -bar{scintillator} +Hneg{layer[-1]} -S1neg{layer[-1]}" +
                    f"|CHol{scintillator} +S1neg{layer[-1]} -S2neg{layer[-1]}" +
                    f"|Void{scintillator} -Shm{scintillator} +Hneg{layer[-1]} -S2neg{layer[-1]}" +
                    f"|Void{scintillator} +S2neg{layer[-1]}")
                
            # Layer 1, 3, and 4
            else:
                f.write(f"\nCLPOS{scintillator}      5 +Clpos{scintillator} -ELpos{layer[-1]}")
                f.write(f"\nCLNEG{scintillator}      5 +Clneg{scintillator} +ELneg{layer[-1]}")
                f.write(f"\nGLPOS{scintillator}      5 +Clpos{scintillator} -Glpos{scintillator} +ELpos{layer[-1]}")
                f.write(f"\nGLNEG{scintillator}      5 +Clneg{scintillator} -Glneg{scintillator} -ELneg{layer[-1]}")
                f.write(f"\nVACPOS{scintillator}      5 +Glpos{scintillator} +ELpos{layer[-1]}")
                f.write(f"\nVACNEG{scintillator}      5 +Glneg{scintillator} -ELneg{layer[-1]}")

            
                f.write(f"\nVOID{scintillator}      5" + 
                        f" +HIn{scintillator} -bar{scintillator} -Hneg{layer[-1]} +Hpos{layer[-1]}" + 
                        f"|+Void{scintillator} -HOut{scintillator} -Hneg{layer[-1]} +Hpos{layer[-1]}" +
                        f"|+SHolp{scintillator} -bar{scintillator} -Hpos{layer[-1]} +S1pos{layer[-1]}" +
                        f"|+CHol{scintillator} -Clpos{scintillator} -S1pos{layer[-1]} +S2pos{layer[-1]}" +
                        f"|Void{scintillator} -Shm{scintillator} -Hpos{layer[-1]} + S2pos{layer[-1]}" +
                        f"|Void{scintillator} -S2pos{layer[-1]} -Clpos{scintillator}" +
                        f"|SHoln{scintillator} -bar{scintillator} +Hneg{layer[-1]} -S1neg{layer[-1]}" +
                        f"|CHol{scintillator} -Clneg{scintillator} +S1neg{layer[-1]} -S2neg{layer[-1]}" +
                        f"|Void{scintillator} -Shm{scintillator} +Hneg{layer[-1]} -S2neg{layer[-1]}" +
                        f"|Void{scintillator} +S2neg{layer[-1]} -Clneg{scintillator}")
                

    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("END","","","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("GEOEND","","","","","","",""))

    # ------------------------------MATERIALS---------------------------
    f.write("\n* ..+....1....+....2....+....3....+....4....+....5....+....6....+....7..")
    f.write("\n* Black hole")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","BLCKHOLE","BLKBODY","","","","",""))

    f.write("\n* Air subsphere")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR","AIR","","","","",""))

    f.write("\n* Void")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","VACUUM","VOID","","","","",""))

    f.write("\n* Support bars")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM","SBO_P1","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM","SBO_N1","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR","SBI_P1","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR","SBI_N1","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM","SBO_P2","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM","SBO_N2","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR","SBI_P2","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR","SBI_N2","","","","",""))

    for layer, scintillators in layers.items():
        for scintillator in scintillators:
            
            # Assigning correct material to the scintillators. Right now OGS is for even named scintillators, while M600 is for odd ones [EXCEPT LAYER 2]
            if int(scintillator) % 2 == 0:  # 0 for odd numbers, 1 for even numbers
                bar_material = "OGS"
                if layer == "Layer2":
                    bar_material = "M600"
            else:
                bar_material = "M600"
                if layer == "Layer2":
                    bar_material = "OGS"
            
            # Making exceptions for Layer 2
            if layer == "Layer2":
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA",bar_material,f"BAR{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"HOUSE{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"S1POS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"S1NEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"S2POS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"S2NEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR",f"VOID{scintillator}","","","","",""))
            # Layer 1, 3 and 4
            else:
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA",bar_material,f"BAR{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","POLYETHY",f"HOUSE{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","POLYETHY",f"S1POS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","POLYETHY",f"S1NEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","POLYETHY",f"S2POS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","POLYETHY",f"S2NEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"CLPOS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","ALUMINUM",f"CLNEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","GLASS",f"GLPOS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","GLASS",f"GLNEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","VACUUM",f"VACPOS{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","VACUUM",f"VACNEG{scintillator}","","","","",""))
                f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","AIR",f"VOID{scintillator}","","","","",""))

    # -------PMMA phantom/target
    f.write("\n* PMMA phantom/target")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","PMMAnew","CASING","","","","",""))
    for i in range(10):
        f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","PMMAnew",f"INS1.{i + 1}","","","","",""))
    for i in range(10):
        f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("ASSIGNMA","PMMAnew",f"INS2.{i + 1}","","","","",""))

    f.write("\n* M600 compound material")   # From Deliverable 1.2
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("MATERIAL","","","1.207","","","","M600"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("COMPOUND","5.69","HYDROGEN","3.96","CARBON","0.4","NITROGEN","M600"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("COMPOUND","0.85","OXYGEN","","","","","M600"))

    f.write("\n* OGS compound material")    # From Hunter's mNOVOv4.inp PHITS input file
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("MATERIAL","","","1.091","","","","OGS"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("COMPOUND","0.456","HYDROGEN","0.532","CARBON","0.012","SILICON","OGS"))

    f.write("\n* PMMA phantom material")    # From Werner et al. 2019; https://iopscience.iop.org/article/10.1088/1361-6560/ab176d
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("MATERIAL","","","1.18","","","","PMMAnew"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("COMPOUND","5.0","CARBON","8.0","HYDROGEN","2.0","OXYGEN","PMMAnew"))

    f.write("\n* Silicate glass (fused quartz)")   # From Wikipedia: https://en.wikipedia.org/wiki/Fused_quartz#List_of_physical_properties
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("MATERIAL","","","2.203","","","","GLASS"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("COMPOUND","1.0","SILICON","2.0","OXYGEN","","","GLASS"))

    # -------------------------------OUTPUT-----------------------------
    # USERDUMP card
    f.write("\n* mgdraw .txt-file dumping")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USERDUMP","100.0","22.0","0.0","1.","","","USRDUMP"))

    f.write("\n* Dose in PMMA phantom")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN","10.","DOSE","-23.","7.5","7.5",f"{str(round(40 + distance_beam_PMMA - nominal_beam_exit_PMMA_distance, 2))}","PMMADOSE"))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN","-7.5","-7.5",f"{str(round(distance_beam_PMMA - nominal_beam_exit_PMMA_distance, 2))}","60.","60.","80.","&"))

    minx = scintillator_coordinate_dict["43"]["Void"][0]
    maxx = scintillator_coordinate_dict["42"]["Void"][1]

    miny = scintillator_coordinate_dict["32"]["Void"][2]
    maxy = scintillator_coordinate_dict["33"]["Void"][3]

    minz = scintillator_coordinate_dict["10"]["Void"][4]
    maxz = scintillator_coordinate_dict["40"]["Void"][5]

    maxs = np.array([[minx, miny, minz], [maxx, maxy, maxz]])
    maxs = NOVO_to_FLUKA_coordinate_transform(maxs)

  
    f.write("\n* Dose in detector core")
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN","10.","DOSE","-25.",f"{round(maxs[1][0], 2)}", f"{round(maxs[0][1], 2)}",f"{round(maxs[1][2], 2)}","DTCCDOSE"))
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN",f"{round(maxs[0][0], 2)}",f"{round(maxs[1][1], 2)}",f"{round(maxs[0][2], 2)}","50","50","50","&"))

    f.write("\n* Photon fluence in detector core")
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN","10.","PHOTON","-26.",f"{round(maxs[1][0],2)}",f"{round(maxs[0][1], 2)}",f"{round(maxs[1][2], 2)}","DTCCgFLU"))
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN",f"{round(maxs[0][0], 2)}",f"{round(maxs[1][1], 2)}",f"{round(maxs[0][2], 2)}","50","50","50","&"))

    f.write("\n* Neutron fluence in detector core")
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN","10.","NEUTRON","-27.",f"{round(maxs[1][0],2)}",f"{round(maxs[0][1],2)}",f"{round(maxs[1][2], 2)}","DTCCnFLU"))
    f.write("\n* {0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("USRBIN",f"{round(maxs[0][0], 2)}",f"{round(maxs[1][1], 2)}",f"{round(maxs[0][2],2)}","50","50","50","&"))

    # --------------------------INITIALIZATION--------------------------
    f.write("\n* Set the random number seed")
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("RANDOMIZ","1.0",str(random.randrange(1, int(9.E8))),"","","","",""))

    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("START","15000000","","","","","",""))
    f.write("\n{0:<10.8s}{1:>10.10s}{2:>10.10s}{3:>10.10s}{4:>10.10s}{5:>10.10s}{6:>10.10s}{7:<8.8s}".format("STOP","","","","","","",""))

    print(f"FLUKA .inp file created: {inp_file_path + "\\" + inp_file_name}")

